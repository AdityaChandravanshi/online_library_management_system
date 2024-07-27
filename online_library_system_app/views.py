from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Author, Book, BorrowRecord
from django.utils.dateparse import parse_date
import openpyxl
from .forms import AuthorForm, BookForm, BorrowRecordForm
from django.contrib import messages
from django.core.exceptions import ValidationError
import re

def index(request):
    return render(request, 'index.html')

def validate_author_data(name, email, bio):
    errors = {}
    if not re.match(r'^[A-Za-z\s]+$', name):
        errors['name'] = 'Name can only contain letters and spaces.'
        raise ValidationError('Name can only contain letters and spaces.')
    if not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
        errors['email'] = 'Enter a valid email address.'
        raise ValidationError('Enter a valid email address.')
    if not bio:
        errors['bio'] = 'Bio cannot be empty.'
        raise ValidationError('Bio cannot be empty.')
    return errors

def add_author(request):
    if request.method == 'POST':
        form = AuthorForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('name')
            email = form.cleaned_data.get('email')
            bio = form.cleaned_data.get('bio')

            try:
                errors = validate_author_data(name, email, bio)
                if not errors:
                    form.save()
                    messages.success(request, 'Author added successfully!')
                    return redirect('author_list')
                else:
                    for field, error in errors.items():
                        form.add_error(field, error)
                    messages.error(request, 'Error adding author. Please correct the errors below.')
            except ValidationError as e:
                form.add_error(None, e.message)
        else:
            messages.error(request, 'Error adding author, Please all fields are required. Please correct the errors below.')
    else:
        form = AuthorForm()
    return render(request, 'add_author.html', {'form': form})

def validate_book_data(title, genre, published_date):
    errors = {}
    if not title:
        errors['title'] = 'Title is required.'
        raise ValidationError('Title is required.')
    if not genre:
        errors['genre'] = 'Genre is required.'
        raise ValidationError('Genre is required.')
    try:
        parse_date(str(published_date))
    except ValueError:
        errors['published_date'] = 'Enter a valid date in YYYY-MM-DD format.'
        raise ValidationError('Enter a valid date in YYYY-MM-DD format.')
    return errors

def add_book(request):
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            title = form.cleaned_data.get('title')
            genre = form.cleaned_data.get('genre')
            published_date = form.cleaned_data.get('published_date')

            try:
                errors = validate_book_data(title, genre, published_date)
                if not errors:
                    form.save()
                    messages.success(request, 'Book added successfully!')
                    return redirect('book_list')
                else:
                    for field, error in errors.items():
                        form.add_error(field, error)
                    messages.error(request, 'Error adding book. Please correct the errors below.')
            except ValidationError as e:
                form.add_error(None, e.message)
        else:
            messages.error(request, 'Error adding book, Please all fields are required. Please correct the errors below.')
    else:
        form = BookForm()
    authors = Author.objects.all()
    return render(request, 'add_book.html', {'form': form, 'authors': authors})

def validate_borrow_record_data(user_name, borrow_date, return_date):
    errors = {}
    if not re.match(r'^[A-Za-z\s]+$', user_name):
        errors['user_name'] = 'User name can only contain letters and spaces.'
        raise ValidationError('User name can only contain letters and spaces.')
    try:
        parse_date(str(borrow_date))
    except ValueError:
        errors['borrow_date'] = 'Enter a valid borrow date in YYYY-MM-DD format.'
        raise ValidationError('Enter a valid borrow date in YYYY-MM-DD format.')
    try:
        parse_date(str(return_date))
    except ValueError:
        errors['return_date'] = 'Enter a valid return date in YYYY-MM-DD format.'
        raise ValidationError('Enter a valid return date in YYYY-MM-DD format.')
    return errors

def add_borrow_record(request):
    if request.method == 'POST':
        form = BorrowRecordForm(request.POST)
        if form.is_valid():
            user_name = form.cleaned_data.get('user_name')
            borrow_date = form.cleaned_data.get('borrow_date')
            return_date = form.cleaned_data.get('return_date')

            try:
                errors = validate_borrow_record_data(user_name, borrow_date, return_date)
                if not errors:
                    form.save()
                    messages.success(request, 'Borrow record added successfully!')
                    return redirect('borrow_list')
                else:
                    for field, error in errors.items():
                        form.add_error(field, error)
                    messages.error(request, 'Error adding borrow record. Please correct the errors below.')
            except ValidationError as e:
                form.add_error(None, e.message)
        else:
            messages.error(request, 'Error adding borrow record, Please all fields are required. Please correct the errors below.')
    else:
        form = BorrowRecordForm()
    books = Book.objects.all()
    return render(request, 'add_borrow_record.html', {'form': form, 'books': books})

def author_list(request):
    authors = Author.objects.all().order_by('id')
    paginator = Paginator(authors, 5)
    page = request.GET.get('page')

    try:
        authors = paginator.page(page)
    except PageNotAnInteger:
        authors = paginator.page(1)
    except EmptyPage:
        authors = paginator.page(paginator.num_pages)

    context = {'authors': authors}
    return render(request, 'author_list.html', context)

def book_list(request):
    books = Book.objects.all().order_by('id')
    paginator = Paginator(books, 5)
    page = request.GET.get('page')

    try:
        books = paginator.page(page)
    except PageNotAnInteger:
        books = paginator.page(1)
    except EmptyPage:
        books = paginator.page(paginator.num_pages)

    context = {'books': books}
    return render(request, 'book_list.html', context)

def borrow_list(request):
    records = BorrowRecord.objects.all().order_by('id')
    paginator = Paginator(records, 5)
    page = request.GET.get('page')

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        records = paginator.page(1)
    except EmptyPage:
        records = paginator.page(paginator.num_pages)

    context = {'records': records}
    return render(request, 'borrow_list.html', context)

def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'

    wb = openpyxl.Workbook()
    author_ws = wb.active
    author_ws.title = 'Authors'

    books_ws = wb.create_sheet(title='Books')
    borrow_ws = wb.create_sheet(title='Borrow Records')

    # Export Authors
    authors = Author.objects.all()
    author_ws.append(['ID', 'Name', 'Email', 'Bio'])
    for author in authors:
        author_ws.append([author.id, author.name, author.email, author.bio])

    # Export Books
    books = Book.objects.all()
    books_ws.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
    for book in books:
        books_ws.append([book.id, book.title, book.genre, book.published_date, book.author.name])

    # Export Borrow Records
    borrow_records = BorrowRecord.objects.all()
    borrow_ws.append(['ID', 'User Name', 'Book', 'Borrow Date', 'Return Date'])
    for record in borrow_records:
        borrow_ws.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

    wb.save(response)
    return response